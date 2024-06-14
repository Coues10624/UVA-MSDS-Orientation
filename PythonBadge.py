import pandas as pd

# Path to the Excel file
file_path = r"C:\Users\Brennan Thompson\Downloads\MSDS-Orientation-Computer-Survey.xlsx"

from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load the workbook and select the active worksheet
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Extract the "CPU Number of Cores" data
cpu_cores = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
    cpu_cores.append(row[5])  # Assuming "CPU Number of Cores" is the 6th column (index 5)

# Create the histogram
plt.hist(cpu_cores, bins='auto', edgecolor='black')
plt.title('Histogram of CPU Number of Cores')
plt.xlabel('Number of Cores')
plt.ylabel('Frequency')
plt.grid(True)

# Show the plot
plt.show()



